"""
A Clunky Python Library containing classes and methods for conducting ratio analysis on 10-K Financial Statements
    Copyright (C) 2020  Andrew Clark
	Contact: ajc1745@gmail.com
"""

import pandas as pd
import os
from openpyxl import load_workbook
from datetime import datetime



KEYWORDS = {"balance sheet" : ['balance sheet'],
            "income statement" : ['income','operation', 'profit','loss'],
            "stockholder\'s equity" : ['equity', 'retained','capital','stock'],
            "cash flow" : ['cash'],
            }


# nyt_data = directory path to folder containing excel reports


# returns the file names of all financial reports in the given directory.
def get_xlsx_data(path):
    statements = os.listdir(path)
    if statements == []:
        raise Exception("No financial reports found")
    for statement in statements:
        if not statement.endswith('.xlsx'):
            raise Exception('only explict xlsx files allowed')
    return statements


# need to look at title cell 'A1' to find match rather than parsing sheet titles
def findsheet(workbook, keyword):
    sheets = workbook.worksheets
    keywords = KEYWORDS[keyword]

    for sheet in sheets:
        sheet_found = a1_match(sheet, keywords)
        if sheet_found == True:
            return sheet.title

    raise Exception("Couldn't find a " + keyword + " worksheet in the report")

# checks the first cell (title cell) for keyword match
def a1_match(sheet, keywords):

    a1 = sheet['A1'].value.lower()
    if ("parenthetical" or "comprehensive" or "tax") in a1:
        return False
    for kw in keywords:
        if kw in a1:
            return True

# converts the shared formatted date string into a date object
def sec_str2date(dstr):
    try:
        date = datetime.strptime(dstr, '%b. %d, %Y').date()
    except ValueError:
        return dstr

    return date

# removes the "12 Months ended" cell present in income statement worksheet

# The Actual API starts here:

class Company:

    # takes the absolute path to the directory holding all SEC financial reports
    def __init__(self, data_path):
        os.chdir(data_path)
        xlsxs = get_xlsx_data(data_path)
        self.path = data_path
        # sort the financial reports from most recent to oldest
        xlsxs.sort(reverse=True)
        self.get_balance_sheets(xlsxs)
        self.convert_date_cols(self.balance_sheet)
        self.get_income_statements(xlsxs)
        self.convert_date_cols(self.income_statement)
        # Potentially could include cash flows, equity statements
        # self.get_cash_flows(xlsxs)
        # self.get_changes_in_equity(xlsxs)


    # parses present excel files for all balance sheet info, adds to relevant dataframe
    def get_balance_sheets(self, xlsxs):
        counter = 0
        for file in xlsxs:
            wb = load_workbook(file)
            balsheet = findsheet(wb, 'balance sheet')

            path = os.path.join(self.path, file)
            temp_frame = pd.read_excel(path, sheet_name=balsheet)
            # if no dataframe has been imported yet, take all columns
            if (counter == 0):
                self.balance_sheet = temp_frame
            # otherwise, only add 3rd column to dataframe
            else:
                frames = [self.balance_sheet, temp_frame.iloc[:,2]]
                self.balance_sheet = pd.concat(frames, axis = 1)
            counter += 1

    # parses present excel files for all income statement info, adds to relevant dataframe
    def get_income_statements(self, xlsxs):
        counter = 0
        for file in xlsxs:
            wb = load_workbook(file)
            income_sheet = findsheet(wb, 'income statement')

            path = os.path.join(self.path, file)
            temp_frame = pd.read_excel(path, sheet_name=income_sheet, skiprows = [0])
            # if no dataframe has been imported yet, take all columns
            if (counter == 0):
                self.income_statement = temp_frame
            # otherwise, only add 3rd column to dataframe
            else:
                frames = [self.income_statement, temp_frame.iloc[:,2]]
                self.income_statement = pd.concat(frames, axis = 1)

            counter += 1



    # Transfers date string column titles into date objects
    def convert_date_cols(self, df):
        df.rename(columns = sec_str2date, inplace = True)


    # todo: Clean up the date headers of columns
    # todo: clean up the row titles
    # todo: figure out what to do with category rows (e.g. current assets)
